
import io
import zipfile
from typing import Dict, List

import pandas as pd
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.worksheet.datavalidation import DataValidation


STATUS_OPTIONS = [
    "Dodáno",
    "Částečně dodáno",
    "Chybí",
    "Není k dispozici",
    "Irelevantní",
]

RELEVANCE_EXPLANATION = {
    "P": "Povinné",
    "D": "Doporučené",
    "V": "Volitelné",
    "-": "Irelevantní",
}


BLOCK_DEFINITIONS = {
    1: "Zadání a organizace",
    2: "Fotodokumentace",
    3: "Energetická data",
    4: "Stavební a objektové technické podklady",
    5: "Technické podklady technologií a energetických zařízení",
    6: "Provozní informace",
    7: "Ekonomika, finance a dotační souvislosti",
}




SERVICE_ROOT_NAMES = {
    "EA": "DPU_EA",
    "PEA": "DPU_PLAN_EA",
    "SPU": "DPU_SPU",
    "EP": "DPU_EP",
    "PENB": "DPU_PENB",
    "ES": "DPU_ES",
    "FVE": "DPU_FVE",
    "KE": "DPU_KE",
    "EM": "DPU_EM",
    "MEK": "DPU_MEK",
    "SP": "DPU_SP",
    "DB": "DPU_DB",
}

ITEM_FOLDER_SHORT_NAMES = {
    "1": "1.1_Smlouva_zadani",
    "2": "1.2_Prehled_objektu",
    "3": "1.3_Schvalovani_terminy",
    "4": "1.4_VZ_pravidla",
    "5": "1.5_Harmonogram_rozhodnuti",
    "6": "2.1_Fotodokumentace",
    "7": "3.1_Faktury_ele",
    "8": "3.2_Smlouvy_ele",
    "9": "3.3_15min_ele",
    "10": "3.4_EAN_parametry",
    "11": "3.5_Faktury_plyn",
    "12": "3.6_Smlouvy_plyn",
    "13": "3.7_Odecty_plyn",
    "14": "3.8_Faktury_teplo",
    "15": "3.9_Smlouvy_teplo",
    "16": "3.10_Odecty_teplo",
    "17": "3.11_Faktury_voda",
    "18": "3.12_Mapovani_OM",
    "19": "3.13_Vyroba_ele",
    "20": "4.1_Dok_stav",
    "21": "4.2_Dok_navrh",
    "22": "4.3_PENB_audity",
    "23": "4.4_Obalka_zatepleni",
    "24": "4.5_Okna_dvere",
    "25": "4.6_Strechy",
    "26": "4.7_Statika_strechy",
    "27": "4.8_Stav_poruchy",
    "28": "5.1_Vytapeni",
    "29": "5.2_Chlazeni",
    "30": "5.3_VZT",
    "31": "5.4_TUV",
    "32": "5.5_Zalozni_zdroje",
    "33": "5.6_FVE_zdroje",
    "34": "5.7_Trafostanice",
    "35": "5.8_MaR_EnMS",
    "36": "5.9_Voda",
    "37": "5.10_Osvetleni",
    "38": "5.11_PBR",
    "39": "5.12_Spotrebice",
    "40": "6.1_Provozni_rezim",
    "41": "6.2_Obsazenost_vyuziti",
    "42": "6.3_Provozni_problemy",
    "43": "6.4_Plan_zmen",
    "44": "6.5_Doplnujici_provoz",
    "45": "7.1_Historie_investic",
    "46": "7.2_Rozpocty_finance",
    "47": "7.3_Dotace",
}

BLOCK_README_TEXTS = {
    1: (
        "Do této části patří zejména smluvní a zadávací dokumenty, přehled řešených objektů, "
        "interní schvalovací proces, pravidla veřejných zakázek a harmonogram projektu."
    ),
    2: (
        "Do této části patří fotodokumentace objektů, technických místností, střech, technologií "
        "a problematických míst."
    ),
    3: (
        "Do této části patří energetická data, zejména faktury, smlouvy, průběhová data, odečty "
        "a přehledy odběrných míst."
    ),
    4: (
        "Do této části patří stavební a objektové technické podklady, například dokumentace stávajícího "
        "a plánovaného stavu, PENB, obvodový plášť, střechy a statické posudky."
    ),
    5: (
        "Do této části patří technické podklady technologií a energetických zařízení, například vytápění, "
        "chlazení, vzduchotechnika, příprava teplé vody, FVE, trafostanice, MaR / EnMS a osvětlení."
    ),
    6: (
        "Do této části patří informace o reálném provozu objektu, jeho obsazenosti, provozních problémech "
        "a plánovaných změnách."
    ),
    7: (
        "Do této části patří informace o investicích, rozpočtových možnostech a případných dotacích "
        "nebo podpůrných programech."
    ),
}


MASTER_ITEMS: List[Dict[str, str | int]] = [
    {
        "id": "1",
        "block": 1,
        "item_no": 1,
        "name": "Smluvní a zadávací dokumenty",
        "description": "Smlouva, objednávka, nabídka, zápis z úvodního jednání a další dokumenty, ze kterých je patrné zadání služby, očekávaný výstup, rozsah a cíle.",
    },
    {
        "id": "2",
        "block": 1,
        "item_no": 2,
        "name": "Přehled řešených objektů / lokalit",
        "description": "Vyplněný soubor PREHLED_OBJEKTU.xlsx se seznamem řešených objektů nebo lokalit ve schválené struktuře.",
    },
    {
        "id": "3",
        "block": 1,
        "item_no": 3,
        "name": "Schvalovací proces a interní termíny",
        "description": "Termíny rady, zastupitelstva, interní schvalovací kroky, kompetence jednotlivých orgánů a vazba rozhodování na projekt.",
    },
    {
        "id": "4",
        "block": 1,
        "item_no": 4,
        "name": "Pravidla veřejných zakázek",
        "description": "Směrnice o zadávání veřejných zakázek, limity pro přímý nákup, poptávku, veřejné výběrové řízení a související kompetence.",
    },
    {
        "id": "5",
        "block": 1,
        "item_no": 5,
        "name": "Harmonogram projektu a klíčová rozhodnutí",
        "description": "Harmonogram zakázky, návazné termíny a zásadní rozhodnutí v průběhu projektu.",
    },
    {
        "id": "6",
        "block": 2,
        "item_no": 1,
        "name": "Fotodokumentace",
        "description": "Fotografie objektů, technických místností, střech, technologií, vad a problematických míst.",
    },
    {
        "id": "7",
        "block": 3,
        "item_no": 1,
        "name": "Faktury za elektřinu",
        "description": "Faktury za dodávku a distribuci elektřiny za jednotlivá odběrná místa nebo objekty, ideálně za posledních 24–36 měsíců.",
    },
    {
        "id": "8",
        "block": 3,
        "item_no": 2,
        "name": "Smlouvy a ceníky elektřiny",
        "description": "Smlouvy s dodavatelem elektřiny, ceníky, dodatky a informace o produktu nebo cenovém mechanismu.",
    },
    {
        "id": "9",
        "block": 3,
        "item_no": 3,
        "name": "15min profily elektřiny",
        "description": "Průběhová data spotřeby nebo výroby elektřiny v 15min kroku, zejména pro hlavní odběry a výrobny.",
    },
    {
        "id": "10",
        "block": 3,
        "item_no": 4,
        "name": "Seznam EAN a parametry odběrných míst elektřiny",
        "description": "Seznam EAN, zatřídění VN / NN, distribuční sazba, velikost jištění, rezervovaný příkon nebo kapacita a další relevantní parametry odběrných míst.",
    },
    {
        "id": "11",
        "block": 3,
        "item_no": 5,
        "name": "Faktury za zemní plyn",
        "description": "Faktury za dodávku a distribuci zemního plynu, ideálně za posledních 24–36 měsíců.",
    },
    {
        "id": "12",
        "block": 3,
        "item_no": 6,
        "name": "Smlouvy na dodávku plynu",
        "description": "Smlouvy s dodavatelem plynu, dodatky a související ceníkové nebo produktové informace.",
    },
    {
        "id": "13",
        "block": 3,
        "item_no": 7,
        "name": "Odečty plynu / distribuční portál",
        "description": "Odečty plynu, exporty z distribučního portálu nebo jiná dostupná provozní data k odběru plynu.",
    },
    {
        "id": "14",
        "block": 3,
        "item_no": 8,
        "name": "Faktury za teplo",
        "description": "Faktury za teplo nebo CZT, ideálně za posledních 24–36 měsíců.",
    },
    {
        "id": "15",
        "block": 3,
        "item_no": 9,
        "name": "Smlouvy na dodávku tepla",
        "description": "Smlouvy o dodávce tepla nebo CZT a související smluvní podmínky.",
    },
    {
        "id": "16",
        "block": 3,
        "item_no": 10,
        "name": "Odečty tepla / distribuční portál",
        "description": "Odečty tepla, exporty z portálu dodavatele nebo jiná dostupná data k odběru tepla.",
    },
    {
        "id": "17",
        "block": 3,
        "item_no": 11,
        "name": "Faktury za vodu",
        "description": "Faktury za vodu a stočné za relevantní objekty.",
    },
    {
        "id": "18",
        "block": 3,
        "item_no": 12,
        "name": "Mapování odběrných míst k objektům",
        "description": "Přehledová tabulka, ve které je každé odběrné místo jednoznačně přiřazeno ke konkrétnímu objektu.",
    },
    {
        "id": "19",
        "block": 3,
        "item_no": 13,
        "name": "Vlastní výroba elektřiny – hodinová nebo průběhová data",
        "description": "Data o vlastní výrobě elektřiny, zejména z FVE nebo jiných vlastních zdrojů, ideálně v hodinovém nebo jiném průběhovém kroku.",
    },
    {
        "id": "20",
        "block": 4,
        "item_no": 1,
        "name": "Dokumentace stávajícího stavu objektů",
        "description": "Pasporty, projektová dokumentace skutečného stavu, technické zprávy, půdorysy, řezy, pohledy a další podklady ke stávajícímu stavu objektu.",
    },
    {
        "id": "21",
        "block": 4,
        "item_no": 2,
        "name": "Projektová dokumentace plánovaného stavu",
        "description": "Studie, projektová dokumentace připravovaných úprav, návrhy rekonstrukcí a další podklady k plánovanému stavu.",
    },
    {
        "id": "22",
        "block": 4,
        "item_no": 3,
        "name": "PENB, energetické audity a další energetické dokumenty",
        "description": "PENB, starší energetické audity, energetické posudky a další obdobné energetické dokumenty.",
    },
    {
        "id": "23",
        "block": 4,
        "item_no": 4,
        "name": "Obvodový plášť a zateplení",
        "description": "Podklady k fasádě, zateplení, skladbám konstrukcí a realizovaným úpravám obvodového pláště.",
    },
    {
        "id": "24",
        "block": 4,
        "item_no": 5,
        "name": "Výplně otvorů",
        "description": "Podklady k oknům, dveřím, vratům a jejich základním parametrům, stáří nebo rozsahu výměn.",
    },
    {
        "id": "25",
        "block": 4,
        "item_no": 6,
        "name": "Střechy a střešní konstrukce",
        "description": "Podklady k typu střechy, skladbě, stavu, ploše a omezením případných zásahů.",
    },
    {
        "id": "26",
        "block": 4,
        "item_no": 7,
        "name": "Statické posouzení střechy",
        "description": "Statické posudky a další podklady důležité pro zásahy do střechy, zejména pro FVE nebo jiné dodatečné zatížení.",
    },
    {
        "id": "27",
        "block": 4,
        "item_no": 8,
        "name": "Stavební stav a poruchy",
        "description": "Informace o vadách, poruchách, zatékání, vlhkosti nebo jiných známých problémech stavební části objektu.",
    },
    {
        "id": "28",
        "block": 5,
        "item_no": 1,
        "name": "Vytápění",
        "description": "Zdroje tepla, rozvody, schémata, regulace, servis, revize a podklady k předávací stanici tepla.",
    },
    {
        "id": "29",
        "block": 5,
        "item_no": 2,
        "name": "Chlazení",
        "description": "Zdroje chladu, chladicí zařízení, technické listy, servis, revize a provozní problémy chlazení.",
    },
    {
        "id": "30",
        "block": 5,
        "item_no": 3,
        "name": "Vzduchotechnika",
        "description": "Vzduchotechnické jednotky, technické listy, servis, revize a provozní problémy VZT.",
    },
    {
        "id": "31",
        "block": 5,
        "item_no": 4,
        "name": "Příprava teplé vody",
        "description": "Podklady ke zdrojům TUV, zásobníkům, cirkulaci, měření a provozním režimům přípravy teplé vody.",
    },
    {
        "id": "32",
        "block": 5,
        "item_no": 5,
        "name": "Záložní zdroje",
        "description": "Dieselagregáty, UPS nebo jiné záložní zdroje včetně základních parametrů a provozních informací.",
    },
    {
        "id": "33",
        "block": 5,
        "item_no": 6,
        "name": "FVE a vlastní zdroje elektřiny",
        "description": "Technické podklady k FVE a dalším vlastním zdrojům elektřiny, včetně smlouvy o připojení výrobny a jednopólového schématu, pokud existují.",
    },
    {
        "id": "34",
        "block": 5,
        "item_no": 7,
        "name": "Trafostanice",
        "description": "Technické podklady, revize a základní parametry trafostanic a souvisejících zařízení.",
    },
    {
        "id": "35",
        "block": 5,
        "item_no": 8,
        "name": "MaR / EnMS",
        "description": "Podklady k měření a regulaci nebo energetickému managementu, zejména popis systému, exporty, trendy, seznam měřených veličin a přístupy.",
    },
    {
        "id": "36",
        "block": 5,
        "item_no": 9,
        "name": "Voda",
        "description": "Technický stav rozvodů vody, úniky, poruchy a další technické podklady k vodnímu hospodářství objektu.",
    },
    {
        "id": "37",
        "block": 5,
        "item_no": 10,
        "name": "Osvětlení",
        "description": "Přehled svítidel a základní podklady k systému osvětlení objektu.",
    },
    {
        "id": "38",
        "block": 5,
        "item_no": 11,
        "name": "PBŘ",
        "description": "Požárně-bezpečnostní řešení objektu a související omezení pro technologie nebo stavební zásahy.",
    },
    {
        "id": "39",
        "block": 5,
        "item_no": 12,
        "name": "Významné spotřebiče",
        "description": "Technologické celky nebo zařízení s významnou spotřebou energie, důležité pro energetickou bilanci objektu.",
    },
    {
        "id": "40",
        "block": 6,
        "item_no": 1,
        "name": "Provozní režim objektu",
        "description": "Provozní doby objektu, pracovní dny, víkendy, směnnost, sezónnost a základní režim fungování objektu.",
    },
    {
        "id": "41",
        "block": 6,
        "item_no": 2,
        "name": "Obsazenost a způsob využití",
        "description": "Počet uživatelů, způsob využití objektu, intenzita provozu a případné sezónní výkyvy.",
    },
    {
        "id": "42",
        "block": 6,
        "item_no": 3,
        "name": "Uživatelské a provozní problémy",
        "description": "Stížnosti uživatelů a opakované provozní problémy, například chlad, přehřívání, hluk, zápach, vlhkost nebo průvan.",
    },
    {
        "id": "43",
        "block": 6,
        "item_no": 4,
        "name": "Plánované změny a rekonstrukce",
        "description": "Připravované stavební úpravy, změny využití, stěhování nebo další plánované zásahy s dopadem na objekt.",
    },
    {
        "id": "44",
        "block": 6,
        "item_no": 5,
        "name": "Doplňující provozní informace",
        "description": "Další provozní informace, které mají význam pro pochopení skutečného fungování objektu nebo areálu.",
    },
    {
        "id": "45",
        "block": 7,
        "item_no": 1,
        "name": "Historie investic",
        "description": "Přehled významných investic do objektu, technologií nebo energetických opatření v posledních letech.",
    },
    {
        "id": "46",
        "block": 7,
        "item_no": 2,
        "name": "Rozpočtové možnosti a finanční rámec",
        "description": "Rozpočtové možnosti klienta, investiční rámec, CAPEX / OPEX limity a další finanční omezení.",
    },
    {
        "id": "47",
        "block": 7,
        "item_no": 3,
        "name": "Dotace a podpůrné programy",
        "description": "Relevantní dotační programy, podmínky podpory a případně rozpracované nebo podané žádosti.",
    },
]


SERVICES = {
    "Energetický audit": {"short": "EA"},
    "Plán energetického auditu": {"short": "PEA"},
    "Studie potenciálu energetických úspor": {"short": "SPU"},
    "Energetický posudek": {"short": "EP"},
    "Průkaz energetické náročnosti budovy": {"short": "PENB"},
    "Energetická studie": {"short": "ES"},
    "Technickoekonomická studie FVE a bateriového úložiště": {"short": "FVE"},
    "Studie komunitní energetiky": {"short": "KE"},
    "Energetický management": {"short": "EM"},
    "Místní energetická koncepce": {"short": "MEK"},
    "Studie proveditelnosti": {"short": "SP"},
    "Správce procesu přípravy projektu Design&Build / EPC / D&B": {"short": "DB"},
}


RELEVANCE_MATRIX: Dict[str, Dict[str, str]] = {
    "1": {"EA":"P","PEA":"P","SPU":"P","EP":"P","PENB":"D","ES":"P","FVE":"P","KE":"P","EM":"P","MEK":"P","SP":"P","DB":"P"},
    "2": {"EA":"P","PEA":"P","SPU":"P","EP":"P","PENB":"P","ES":"P","FVE":"P","KE":"P","EM":"D","MEK":"P","SP":"P","DB":"D"},
    "3": {"EA":"D","PEA":"P","SPU":"D","EP":"V","PENB":"-","ES":"D","FVE":"D","KE":"D","EM":"D","MEK":"D","SP":"D","DB":"P"},
    "4": {"EA":"V","PEA":"D","SPU":"V","EP":"-","PENB":"-","ES":"D","FVE":"D","KE":"V","EM":"V","MEK":"V","SP":"D","DB":"P"},
    "5": {"EA":"D","PEA":"P","SPU":"D","EP":"V","PENB":"-","ES":"D","FVE":"D","KE":"D","EM":"D","MEK":"D","SP":"P","DB":"P"},
    "6": {"EA":"V","PEA":"V","SPU":"V","EP":"V","PENB":"P","ES":"V","FVE":"P","KE":"V","EM":"V","MEK":"V","SP":"V","DB":"V"},
    "7": {"EA":"P","PEA":"P","SPU":"P","EP":"P","PENB":"-","ES":"P","FVE":"P","KE":"P","EM":"P","MEK":"D","SP":"D","DB":"-"},
    "8": {"EA":"D","PEA":"D","SPU":"D","EP":"D","PENB":"-","ES":"P","FVE":"P","KE":"P","EM":"D","MEK":"D","SP":"D","DB":"-"},
    "9": {"EA":"D","PEA":"D","SPU":"D","EP":"D","PENB":"-","ES":"P","FVE":"P","KE":"P","EM":"D","MEK":"V","SP":"V","DB":"-"},
    "10": {"EA":"P","PEA":"P","SPU":"P","EP":"P","PENB":"-","ES":"P","FVE":"P","KE":"P","EM":"P","MEK":"D","SP":"D","DB":"-"},
    "11": {"EA":"P","PEA":"P","SPU":"P","EP":"D","PENB":"-","ES":"P","FVE":"-","KE":"-","EM":"D","MEK":"D","SP":"D","DB":"-"},
    "12": {"EA":"D","PEA":"D","SPU":"D","EP":"D","PENB":"-","ES":"P","FVE":"-","KE":"-","EM":"D","MEK":"D","SP":"D","DB":"-"},
    "13": {"EA":"D","PEA":"D","SPU":"D","EP":"D","PENB":"-","ES":"D","FVE":"-","KE":"-","EM":"D","MEK":"V","SP":"V","DB":"-"},
    "14": {"EA":"P","PEA":"P","SPU":"P","EP":"D","PENB":"-","ES":"P","FVE":"-","KE":"-","EM":"D","MEK":"D","SP":"D","DB":"-"},
    "15": {"EA":"D","PEA":"D","SPU":"D","EP":"D","PENB":"-","ES":"P","FVE":"-","KE":"-","EM":"D","MEK":"D","SP":"D","DB":"-"},
    "16": {"EA":"D","PEA":"D","SPU":"D","EP":"D","PENB":"-","ES":"D","FVE":"-","KE":"-","EM":"D","MEK":"V","SP":"V","DB":"-"},
    "17": {"EA":"D","PEA":"D","SPU":"D","EP":"-","PENB":"-","ES":"D","FVE":"-","KE":"-","EM":"D","MEK":"V","SP":"V","DB":"-"},
    "18": {"EA":"P","PEA":"P","SPU":"P","EP":"P","PENB":"-","ES":"P","FVE":"P","KE":"P","EM":"P","MEK":"D","SP":"D","DB":"-"},
    "19": {"EA":"V","PEA":"V","SPU":"V","EP":"V","PENB":"-","ES":"D","FVE":"P","KE":"P","EM":"D","MEK":"D","SP":"D","DB":"-"},
    "20": {"EA":"P","PEA":"P","SPU":"P","EP":"D","PENB":"P","ES":"P","FVE":"D","KE":"D","EM":"V","MEK":"D","SP":"D","DB":"D"},
    "21": {"EA":"D","PEA":"P","SPU":"D","EP":"D","PENB":"-","ES":"P","FVE":"P","KE":"D","EM":"V","MEK":"D","SP":"P","DB":"P"},
    "22": {"EA":"D","PEA":"P","SPU":"P","EP":"D","PENB":"P","ES":"P","FVE":"D","KE":"D","EM":"D","MEK":"D","SP":"D","DB":"V"},
    "23": {"EA":"P","PEA":"P","SPU":"P","EP":"D","PENB":"P","ES":"P","FVE":"V","KE":"V","EM":"-","MEK":"D","SP":"D","DB":"-"},
    "24": {"EA":"P","PEA":"P","SPU":"P","EP":"D","PENB":"P","ES":"P","FVE":"V","KE":"V","EM":"-","MEK":"D","SP":"D","DB":"-"},
    "25": {"EA":"P","PEA":"P","SPU":"P","EP":"D","PENB":"P","ES":"P","FVE":"P","KE":"D","EM":"-","MEK":"D","SP":"D","DB":"-"},
    "26": {"EA":"V","PEA":"V","SPU":"V","EP":"-","PENB":"-","ES":"D","FVE":"P","KE":"D","EM":"-","MEK":"V","SP":"D","DB":"-"},
    "27": {"EA":"P","PEA":"P","SPU":"P","EP":"D","PENB":"D","ES":"P","FVE":"D","KE":"D","EM":"V","MEK":"D","SP":"D","DB":"D"},
    "28": {"EA":"P","PEA":"P","SPU":"P","EP":"P","PENB":"D","ES":"P","FVE":"-","KE":"-","EM":"D","MEK":"D","SP":"D","DB":"D"},
    "29": {"EA":"D","PEA":"D","SPU":"D","EP":"D","PENB":"V","ES":"P","FVE":"-","KE":"-","EM":"V","MEK":"V","SP":"D","DB":"D"},
    "30": {"EA":"P","PEA":"P","SPU":"P","EP":"D","PENB":"D","ES":"P","FVE":"-","KE":"-","EM":"V","MEK":"V","SP":"D","DB":"D"},
    "31": {"EA":"D","PEA":"D","SPU":"D","EP":"D","PENB":"D","ES":"D","FVE":"-","KE":"-","EM":"V","MEK":"V","SP":"D","DB":"-"},
    "32": {"EA":"V","PEA":"V","SPU":"V","EP":"V","PENB":"-","ES":"D","FVE":"D","KE":"D","EM":"V","MEK":"V","SP":"D","DB":"D"},
    "33": {"EA":"V","PEA":"V","SPU":"V","EP":"V","PENB":"-","ES":"D","FVE":"P","KE":"P","EM":"D","MEK":"D","SP":"D","DB":"D"},
    "34": {"EA":"V","PEA":"V","SPU":"V","EP":"V","PENB":"-","ES":"D","FVE":"P","KE":"P","EM":"V","MEK":"V","SP":"D","DB":"D"},
    "35": {"EA":"D","PEA":"D","SPU":"D","EP":"D","PENB":"-","ES":"P","FVE":"D","KE":"D","EM":"P","MEK":"D","SP":"D","DB":"D"},
    "36": {"EA":"V","PEA":"V","SPU":"V","EP":"-","PENB":"-","ES":"D","FVE":"-","KE":"-","EM":"D","MEK":"V","SP":"V","DB":"-"},
    "37": {"EA":"D","PEA":"D","SPU":"D","EP":"-","PENB":"-","ES":"D","FVE":"-","KE":"-","EM":"V","MEK":"V","SP":"D","DB":"-"},
    "38": {"EA":"V","PEA":"V","SPU":"V","EP":"V","PENB":"-","ES":"D","FVE":"P","KE":"D","EM":"-","MEK":"-","SP":"D","DB":"D"},
    "39": {"EA":"D","PEA":"D","SPU":"D","EP":"D","PENB":"-","ES":"P","FVE":"P","KE":"P","EM":"D","MEK":"D","SP":"D","DB":"D"},
    "40": {"EA":"P","PEA":"P","SPU":"P","EP":"D","PENB":"D","ES":"P","FVE":"D","KE":"P","EM":"P","MEK":"D","SP":"D","DB":"V"},
    "41": {"EA":"P","PEA":"P","SPU":"P","EP":"D","PENB":"P","ES":"P","FVE":"D","KE":"P","EM":"P","MEK":"D","SP":"D","DB":"V"},
    "42": {"EA":"D","PEA":"D","SPU":"D","EP":"D","PENB":"D","ES":"P","FVE":"V","KE":"D","EM":"D","MEK":"V","SP":"D","DB":"V"},
    "43": {"EA":"D","PEA":"P","SPU":"D","EP":"D","PENB":"-","ES":"P","FVE":"D","KE":"D","EM":"D","MEK":"D","SP":"P","DB":"P"},
    "44": {"EA":"V","PEA":"V","SPU":"V","EP":"V","PENB":"-","ES":"D","FVE":"V","KE":"D","EM":"D","MEK":"V","SP":"D","DB":"V"},
    "45": {"EA":"D","PEA":"P","SPU":"P","EP":"D","PENB":"-","ES":"P","FVE":"D","KE":"D","EM":"D","MEK":"D","SP":"P","DB":"P"},
    "46": {"EA":"D","PEA":"P","SPU":"P","EP":"D","PENB":"-","ES":"P","FVE":"P","KE":"D","EM":"D","MEK":"D","SP":"P","DB":"P"},
    "47": {"EA":"V","PEA":"D","SPU":"D","EP":"V","PENB":"-","ES":"D","FVE":"D","KE":"D","EM":"D","MEK":"D","SP":"P","DB":"D"},
}


ROOT_README_TEXT = """NÁZEV: Hlavní složka projektu – podklady pro službu DPU ENERGY

K ČEMU TATO SLOŽKA SLOUŽÍ:
Tato složka slouží jako centrální úložiště podkladů pro zpracování vybrané služby.

JAK SLOŽKU POUŽÍVAT:
- Soubory nahrávejte do tematicky odpovídajících hlavních složek 01 až 07.
- Pokud si nejste jistí, kam dokument patří, uložte jej do složky 99_ARCHIV_NEZARAZENO.
- Názvy hlavních složek prosím neměňte.
- Při pojmenování souborů používejte pravidla uvedená v souboru KLIC_POJMENOVANI_SOUBORU.txt.
"""

NAMING_KEY_TEXT = """KLÍČ POJMENOVÁNÍ SOUBORŮ

OBECNÝ FORMÁT:
[OBLAST]_[OBJEKT nebo ODBĚR]_[POPIS]_[OBDOBÍ nebo DATUM]_[VERZE]

DOPORUČENÍ:
- používat pouze písmena bez diakritiky, čísla a podtržítka
- nepoužívat mezery
- datum uvádět ve formátu YYYY-MM nebo YYYY-MM-DD
- verzi uvádět například v01, v02
- pokud není znám objekt, použijte identifikátor odběrného místa nebo obecný popis
"""

BLOCK_FOLDERS = {
    1: "01_ZADANI",
    2: "02_FOTO",
    3: "03_ENERG_DATA",
    4: "04_STAVEBNI",
    5: "05_TECHNOLOGIE",
    6: "06_PROVOZ",
    7: "07_EKONOMIKA",
}

BLOCK_READMES = {
    1: """NÁZEV: 01_ZADANI

K ČEMU SLOŽKA SLOUŽÍ:
Do této složky patří dokumenty, ze kterých je patrné zadání služby, rozsah řešených objektů, interní schvalovací proces, pravidla veřejných zakázek a harmonogram projektu.

CO TYPICKY NAHRÁT:
- smlouvu, objednávku nebo nabídku
- zápis z úvodního jednání
- přehled řešených objektů / lokalit
- termíny rady a zastupitelstva, pokud mají vazbu na projekt
- interní pravidla veřejných zakázek
- harmonogram projektu

POZNÁMKA:
Kontaktní osoby k jednotlivým objektům vyplňte primárně do souboru PREHLED_OBJEKTU.xlsx.
""",
    2: """NÁZEV: 02_FOTO

K ČEMU SLOŽKA SLOUŽÍ:
Do této složky patří fotodokumentace objektů, technických místností, střech, technologií a problematických míst.

CO TYPICKY NAHRÁT:
- fotografie exteriéru objektu
- fotografie technických místností
- fotografie střech
- fotografie vad, poruch nebo problematických detailů
- fotografie technologií
""",
    3: """NÁZEV: 03_ENERG_DATA

K ČEMU SLOŽKA SLOUŽÍ:
Do této složky patří energetická data, zejména faktury, smlouvy, průběhová data, odečty a přehledy odběrných míst.

CO TYPICKY NAHRÁT:
- faktury za elektřinu, zemní plyn, teplo a vodu
- smlouvy a ceníky elektřiny
- smlouvy na dodávku plynu nebo tepla
- 15min profily elektřiny
- odečty z distribučních portálů
- seznam EAN a parametry odběrných míst
- mapování odběrných míst k objektům
- data o vlastní výrobě elektřiny
""",
    4: """NÁZEV: 04_STAVEBNI

K ČEMU SLOŽKA SLOUŽÍ:
Do této složky patří stavební a objektové technické podklady ke stávajícímu a případně i plánovanému stavu objektu.

CO TYPICKY NAHRÁT:
- pasporty objektů
- projektovou dokumentaci stávajícího stavu
- projektovou dokumentaci plánovaného stavu
- PENB, energetické audity a další energetické dokumenty
- podklady k obvodovému plášti, zateplení a výplním otvorů
- podklady ke střechám a střešním konstrukcím
- statické posouzení střechy
- informace o stavebním stavu a známých poruchách
""",
    5: """NÁZEV: 05_TECHNOLOGIE

K ČEMU SLOŽKA SLOUŽÍ:
Do této složky patří technické podklady k vytápění, chlazení, vzduchotechnice, přípravě teplé vody, FVE, trafostanicím, MaR / EnMS a dalším technologiím.

CO TYPICKY NAHRÁT:
- podklady k vytápění a zdrojům tepla
- podklady k chlazení
- podklady ke vzduchotechnice
- podklady k přípravě teplé vody
- podklady k záložním zdrojům
- podklady k FVE a vlastním zdrojům elektřiny
- podklady k trafostanicím
- podklady k MaR / EnMS
- technické podklady k vodě, osvětlení nebo PBŘ
- přehled významných spotřebičů
""",
    6: """NÁZEV: 06_PROVOZ

K ČEMU SLOŽKA SLOUŽÍ:
Do této složky patří informace o reálném provozu objektu, jeho obsazenosti, uživatelských problémech a plánovaných změnách.

CO TYPICKY NAHRÁT:
- informace o provozních dobách a režimu objektu
- informace o obsazenosti a způsobu využití
- přehled uživatelských nebo provozních problémů
- informace o plánovaných změnách a rekonstrukcích
- další doplňující provozní informace
""",
    7: """NÁZEV: 07_EKONOMIKA

K ČEMU SLOŽKA SLOUŽÍ:
Do této složky patří informace o historii investic, finančních možnostech klienta a případných dotacích nebo podpůrných programech.

CO TYPICKY NAHRÁT:
- přehled minulých investic
- informace o rozpočtových možnostech a finančním rámci
- informace o relevantních dotačních programech
- rozpracované nebo podané dotační žádosti, pokud existují
""",
}



def format_readme_for_ui(text: str) -> str:
    return text.replace("\r\n", "\n").strip()


def parse_query_params() -> Dict[str, str]:
    params = st.query_params
    result = {}
    for key in ["customerName", "projectCode", "projectName"]:
        value = params.get(key, "")
        if isinstance(value, list):
            result[key] = value[0] if value else ""
        else:
            result[key] = value
    return result


def safe_name(text: str) -> str:
    bad = '<>:"/\\|?*'
    out = "".join("_" if ch in bad else ch for ch in str(text).strip())
    out = out.replace(" ", "_")
    return out or "PROJEKT"


def normalize_text(text: str) -> str:
    import unicodedata
    text = "" if text is None else str(text)
    text = unicodedata.normalize("NFKD", text)
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    return text.strip().lower()

def get_service_root_folder(service_name: str) -> str:
    service_short = SERVICES[service_name]["short"]
    return SERVICE_ROOT_NAMES.get(service_short, f"DPU_{service_short}")


def get_item_folder_name(item_id: str) -> str:
    return ITEM_FOLDER_SHORT_NAMES.get(str(item_id), safe_name(str(item_id)))


def shorten_object_folder_name(obj_no: str, obj_label: str, max_label_length: int = 28) -> str:
    cleaned = safe_name(obj_label)
    if len(cleaned) > max_label_length:
        cleaned = cleaned[:max_label_length].rstrip("_")
    if obj_no:
        return f"{obj_no}_{cleaned}" if cleaned else str(obj_no)
    return cleaned or "OBJEKT"



def dataframe_to_excel_bytes(
    df: pd.DataFrame,
    sheet_name: str,
    dropdowns: Dict[str, List[str]] | None = None,
    freeze_cell: str = "A2",
) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    for c_idx, col in enumerate(df.columns, start=1):
        cell = ws.cell(row=1, column=c_idx, value=col)
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9E2F3")
        cell.alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")

    for r_idx, row in enumerate(df.itertuples(index=False), start=2):
        for c_idx, value in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx, value=value)

    ws.freeze_panes = freeze_cell
    ws.auto_filter.ref = ws.dimensions
    ws.row_dimensions[1].height = 36

    for col_cells in ws.columns:
        max_len = 0
        col_letter = col_cells[0].column_letter
        for cell in col_cells:
            value = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(value))
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 12), 45)

    if dropdowns:
        header_map = {cell.value: cell.column for cell in ws[1]}
        for column_name, options in dropdowns.items():
            if column_name not in header_map:
                continue
            col_idx = header_map[column_name]
            col_letter = ws.cell(row=1, column=col_idx).column_letter
            dv = DataValidation(
                type="list",
                formula1=f'"{",".join(options)}"',
                allow_blank=True,
            )
            dv.promptTitle = "Vyberte hodnotu"
            dv.prompt = "Použijte jednu z nabízených možností."
            dv.errorTitle = "Neplatná hodnota"
            dv.error = "Použijte jednu z povolených hodnot."
            ws.add_data_validation(dv)
            dv.add(f"{col_letter}2:{col_letter}1000")

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def build_service_dataframe(service_name: str) -> pd.DataFrame:
    service_short = SERVICES[service_name]["short"]
    rows = []

    for item in MASTER_ITEMS:
        relevance = RELEVANCE_MATRIX.get(str(item["id"]), {}).get(service_short, "-")
        if relevance == "-":
            continue

        row = item.copy()
        row["relevance"] = relevance
        row["relevance_text"] = RELEVANCE_EXPLANATION[relevance]
        row["default_selected"] = True
        row["display_number"] = f"{row['block']}.{row['item_no']}"
        row["block_name"] = BLOCK_DEFINITIONS[row["block"]]
        rows.append(row)

    return pd.DataFrame(rows)


OBJECT_OVERVIEW_HEADERS = [
    "Pořadové číslo objektu",
    "Název objektu",
    "Adresa objektu / parcelní číslo",
    "Vlastník objektu",
    "Provozovatel objektu",
    "Využívaná část objektu (odhad v %)",
    "Jméno a příjmení kontaktu",
    "Funkce",
    "Telefon",
    "E-mail",
    "Majetkoprávní vztah",
    "Památková ochrana / správní omezení",
    "Poznámka",
]


NORMALIZED_OBJECT_COLUMN_MAP = {
    "poradove cislo objektu": "Pořadové číslo objektu",
    "poradove_cislo_objektu": "Pořadové číslo objektu",
    "nazev objektu": "Název objektu",
    "nazev_objektu": "Název objektu",
    "objekt": "Název objektu",
    "adresa objektu / parcelni cislo": "Adresa objektu / parcelní číslo",
    "adresa objektu": "Adresa objektu / parcelní číslo",
    "adresa": "Adresa objektu / parcelní číslo",
    "parcelni cislo": "Adresa objektu / parcelní číslo",
    "vlastnik objektu": "Vlastník objektu",
    "vlastnik": "Vlastník objektu",
    "provozovatel objektu": "Provozovatel objektu",
    "provozovatel": "Provozovatel objektu",
    "vyuzivana cast objektu (odhad v %)": "Využívaná část objektu (odhad v %)",
    "vyuzivana cast objektu": "Využívaná část objektu (odhad v %)",
    "vyuziti v %": "Využívaná část objektu (odhad v %)",
    "jmeno a prijmeni kontaktu": "Jméno a příjmení kontaktu",
    "kontakt": "Jméno a příjmení kontaktu",
    "funkce": "Funkce",
    "telefon": "Telefon",
    "e-mail": "E-mail",
    "email": "E-mail",
    "majetkopravni vztah": "Majetkoprávní vztah",
    "pamatkova ochrana / spravni omezeni": "Památková ochrana / správní omezení",
    "spravni omezeni": "Památková ochrana / správní omezení",
    "poznamka": "Poznámka",
}


def coerce_object_df(df: pd.DataFrame) -> pd.DataFrame:
    out = pd.DataFrame(columns=OBJECT_OVERVIEW_HEADERS)
    normalized_cols = {normalize_text(c): c for c in df.columns}
    for norm_name, target in NORMALIZED_OBJECT_COLUMN_MAP.items():
        if norm_name in normalized_cols:
            out[target] = df[normalized_cols[norm_name]]

    if "Název objektu" not in out.columns or out["Název objektu"].replace("", pd.NA).isna().all():
        cols = list(df.columns)
        if len(cols) >= 1:
            out["Název objektu"] = df[cols[0]]
        if len(cols) >= 2:
            out["Adresa objektu / parcelní číslo"] = df[cols[1]]

    if "Pořadové číslo objektu" not in out.columns or out["Pořadové číslo objektu"].replace("", pd.NA).isna().all():
        out["Pořadové číslo objektu"] = range(1, len(df) + 1)

    for header in OBJECT_OVERVIEW_HEADERS:
        if header not in out.columns:
            out[header] = ""

    out = out[OBJECT_OVERVIEW_HEADERS].copy().fillna("")
    return out


def parse_uploaded_objects(uploaded_file) -> pd.DataFrame:
    if uploaded_file is None:
        return pd.DataFrame(columns=OBJECT_OVERVIEW_HEADERS)
    name = uploaded_file.name.lower()
    if name.endswith(".csv"):
        df = pd.read_csv(uploaded_file)
    else:
        df = pd.read_excel(uploaded_file)
    return coerce_object_df(df)


def parse_manual_objects(text: str) -> pd.DataFrame:
    rows = []
    lines = [line.strip() for line in (text or "").splitlines() if line.strip()]
    for idx, line in enumerate(lines, start=1):
        parts = [p.strip() for p in line.split("|")]
        row = {h: "" for h in OBJECT_OVERVIEW_HEADERS}
        row["Pořadové číslo objektu"] = idx
        row["Název objektu"] = parts[0] if len(parts) > 0 else ""
        row["Adresa objektu / parcelní číslo"] = parts[1] if len(parts) > 1 else ""
        row["Vlastník objektu"] = parts[2] if len(parts) > 2 else ""
        row["Provozovatel objektu"] = parts[3] if len(parts) > 3 else ""
        rows.append(row)
    return pd.DataFrame(rows, columns=OBJECT_OVERVIEW_HEADERS) if rows else pd.DataFrame(columns=OBJECT_OVERVIEW_HEADERS)


def get_object_input_df(uploaded_file, manual_text: str) -> pd.DataFrame:
    uploaded_df = parse_uploaded_objects(uploaded_file) if uploaded_file is not None else pd.DataFrame(columns=OBJECT_OVERVIEW_HEADERS)
    if not uploaded_df.empty:
        return uploaded_df
    manual_df = parse_manual_objects(manual_text)
    if not manual_df.empty:
        return manual_df
    return pd.DataFrame(columns=OBJECT_OVERVIEW_HEADERS)


def build_checklist_xlsx(service_df: pd.DataFrame, service_name: str) -> bytes:
    export_df = service_df.copy()
    export_df["Stav"] = "Chybí"
    export_df["Poznámka"] = ""
    export_df = export_df[["display_number", "block_name", "name", "description", "relevance_text", "Stav", "Poznámka"]].rename(columns={"display_number": "Číslo položky", "block_name": "Blok", "name": "Položka", "description": "Co se očekává", "relevance_text": "Relevance"})
    dropdowns = {"Stav": STATUS_OPTIONS}
    return dataframe_to_excel_bytes(export_df, "Checklist", dropdowns=dropdowns, freeze_cell="A2")


def build_prehled_objektu_xlsx(service_df: pd.DataFrame, object_df: pd.DataFrame | None = None, object_mode: bool = False) -> bytes:
    rows_df = object_df.copy() if object_df is not None and not object_df.empty else pd.DataFrame(columns=OBJECT_OVERVIEW_HEADERS)
    if rows_df.empty:
        rows = []
        for i in range(1, 4):
            row = {h: "" for h in OBJECT_OVERVIEW_HEADERS}
            row["Pořadové číslo objektu"] = i
            rows.append(row)
        rows_df = pd.DataFrame(rows, columns=OBJECT_OVERVIEW_HEADERS)

    relevant_df = service_df[service_df["block"].isin([2, 3, 4, 5, 6])].copy() if object_mode else service_df.copy()
    item_columns = [f"{row['display_number']} | {row['name']}" for _, row in relevant_df.iterrows()]
    export_df = rows_df.copy()
    for col in item_columns:
        if col not in export_df.columns:
            export_df[col] = "Chybí"

    export_df = export_df[OBJECT_OVERVIEW_HEADERS + item_columns]
    dropdown_cols = {col: ["Nahráno", "Chybí", "Není k dispozici", "Irelevantní"] for col in item_columns}
    return dataframe_to_excel_bytes(export_df, "Přehled_objektů", dropdowns=dropdown_cols, freeze_cell="D2")


def build_centralni_prehled_xlsx(service_df: pd.DataFrame, service_name: str) -> bytes:
    central_df = service_df[service_df["block"].isin([1, 7])].copy()
    item_columns = [f"{row['display_number']} | {row['name']}" for _, row in central_df.iterrows()]
    row = {"Služba": service_name, "Název zákazníka": "", "Číslo projektu v Caflou": "", "Název projektu v Caflou": "", "Poznámka": ""}
    for col in item_columns:
        row[col] = "Chybí"
    df = pd.DataFrame([row])
    ordered_cols = ["Služba", "Název zákazníka", "Číslo projektu v Caflou", "Název projektu v Caflou", "Poznámka"] + item_columns
    dropdown_cols = {col: ["Nahráno", "Chybí", "Není k dispozici", "Irelevantní"] for col in item_columns}
    return dataframe_to_excel_bytes(df[ordered_cols], "Centrální_podklady", dropdowns=dropdown_cols, freeze_cell="F2")


OBJECT_MODE_ROOT_README = """NÁZEV: Hlavní složka projektu – podklady pro službu DPU ENERGY

K ČEMU TATO SLOŽKA SLOUŽÍ:
Tato varianta slouží pro přehledné ukládání podkladů po jednotlivých objektech.

JAK SLOŽKU POUŽÍVAT:
- centrální dokumenty nahrávejte do složek 01_ZADANI_A_ORGANIZACE a 02_EKONOMIKA_FINANCE_A_DOTACE
- objektové podklady nahrávejte do složky 03_OBJEKTY, kde má každý objekt svou vlastní podsložku
- uvnitř každého objektu jsou tematické bloky 01 až 05
- pokud si nejste jistí, kam dokument patří, uložte jej do složky 99_ARCHIV_NEZARAZENO
"""

OBJECTS_FOLDER_README = """NÁZEV: 03_OBJEKTY

K ČEMU SLOŽKA SLOUŽÍ:
Tato složka obsahuje jednotlivé objekty. Každý objekt má vlastní podsložku a uvnitř ní bloky 01 až 05 pro objektové podklady.

JAK SLOŽKU POUŽÍVAT:
- pro každý objekt používejte samostatnou složku
- do objektové složky nahrávejte pouze podklady vztahující se ke konkrétnímu objektu
- centrální dokumenty patří mimo tuto složku do 01 nebo 02 na hlavní úrovni
"""

OBJECT_MODE_BLOCK_FOLDERS = {
    2: "01_FOTO",
    3: "02_ENERG_DATA",
    4: "03_STAVEBNI",
    5: "04_TECHNOLOGIE",
    6: "05_PROVOZ",
}


def get_block_intro(block_no: int) -> str:
    raw = format_readme_for_ui(BLOCK_READMES.get(block_no, BLOCK_README_TEXTS.get(block_no, "")))
    lines = [line.strip() for line in raw.splitlines() if line.strip()]
    for idx, line in enumerate(lines):
        if line.startswith("K ČEMU SLOŽKA SLOUŽÍ:") and idx + 1 < len(lines):
            return lines[idx + 1]
    return BLOCK_README_TEXTS.get(block_no, "")


def build_zip_package(customer_name: str, project_code: str, project_name: str, service_name: str, checklist_bytes: bytes, prehled_objektu_bytes: bytes, selected_df: pd.DataFrame, structure_mode: str, centralni_prehled_bytes: bytes | None = None, object_df: pd.DataFrame | None = None) -> bytes:
    project_folder = get_service_root_folder(service_name)

    def write_dir(zf: zipfile.ZipFile, path: str) -> None:
        directory = path if path.endswith("/") else f"{path}/"
        zf.writestr(directory, "")

    output = io.BytesIO()
    with zipfile.ZipFile(output, "w", zipfile.ZIP_DEFLATED) as zf:
        write_dir(zf, project_folder)
        zf.writestr(f"{project_folder}/README.txt", OBJECT_MODE_ROOT_README if structure_mode == "object" else ROOT_README_TEXT)
        zf.writestr(f"{project_folder}/KLIC_POJMENOVANI_SOUBORU.txt", NAMING_KEY_TEXT)

        if structure_mode == "standard":
            for block_no, folder in BLOCK_FOLDERS.items():
                write_dir(zf, f"{project_folder}/{folder}")
                zf.writestr(f"{project_folder}/{folder}/README.txt", BLOCK_READMES[block_no])
            for _, row in selected_df.iterrows():
                block_folder = BLOCK_FOLDERS[int(row["block"])]
                item_folder = get_item_folder_name(str(row["id"]))
                write_dir(zf, f"{project_folder}/{block_folder}/{item_folder}")
            zf.writestr(f"{project_folder}/{BLOCK_FOLDERS[1]}/CHECKLIST_PODKLADU.xlsx", checklist_bytes)
            zf.writestr(f"{project_folder}/{BLOCK_FOLDERS[1]}/PREHLED_OBJEKTU.xlsx", prehled_objektu_bytes)
        else:
            root1 = "01_ZADANI"
            root2 = "02_EKONOMIKA"
            root3 = "03_OBJEKTY"
            write_dir(zf, f"{project_folder}/{root1}")
            write_dir(zf, f"{project_folder}/{root2}")
            write_dir(zf, f"{project_folder}/{root3}")
            zf.writestr(f"{project_folder}/{root1}/README.txt", BLOCK_READMES[1])
            zf.writestr(f"{project_folder}/{root2}/README.txt", BLOCK_READMES[7])
            zf.writestr(f"{project_folder}/{root3}/README.txt", OBJECTS_FOLDER_README)
            for _, row in selected_df[selected_df["block"] == 1].iterrows():
                item_folder = get_item_folder_name(str(row["id"]))
                write_dir(zf, f"{project_folder}/{root1}/{item_folder}")
            for _, row in selected_df[selected_df["block"] == 7].iterrows():
                item_folder = get_item_folder_name(str(row["id"]))
                write_dir(zf, f"{project_folder}/{root2}/{item_folder}")

            object_rows = object_df.copy() if object_df is not None and not object_df.empty else pd.DataFrame([{"Pořadové číslo objektu": 1, "Název objektu": "OBJEKT_01"}])
            object_related = selected_df[selected_df["block"].isin([2, 3, 4, 5, 6])].copy()
            for _, obj in object_rows.iterrows():
                obj_label = str(obj.get("Název objektu", "")).strip() or f"OBJEKT_{obj.get('Pořadové číslo objektu', '')}"
                obj_no = str(obj.get("Pořadové číslo objektu", "")).strip()
                obj_folder_name = shorten_object_folder_name(obj_no, obj_label)
                object_root = f"{project_folder}/{root3}/{obj_folder_name}"
                write_dir(zf, object_root)
                for block_no, folder_name in OBJECT_MODE_BLOCK_FOLDERS.items():
                    block_path = f"{object_root}/{folder_name}"
                    write_dir(zf, block_path)
                    zf.writestr(f"{block_path}/README.txt", BLOCK_READMES[block_no])
                for _, row in object_related.iterrows():
                    subblock = OBJECT_MODE_BLOCK_FOLDERS[int(row["block"])]
                    item_folder = get_item_folder_name(str(row["id"]))
                    write_dir(zf, f"{object_root}/{subblock}/{item_folder}")

            zf.writestr(f"{project_folder}/{root1}/CHECKLIST_PODKLADU.xlsx", checklist_bytes)
            zf.writestr(f"{project_folder}/{root1}/PREHLED_OBJEKTU.xlsx", prehled_objektu_bytes)
            if centralni_prehled_bytes is not None:
                zf.writestr(f"{project_folder}/{root1}/PREHLED_CENTRALNICH_PODKLADU.xlsx", centralni_prehled_bytes)

        write_dir(zf, f"{project_folder}/99_ARCHIV_NEZARAZENO")

    output.seek(0)
    return output.getvalue()


st.set_page_config(page_title="Podklady pro studie DPU ENERGY", layout="wide")
st.markdown("""
<style>
[data-testid="stHeader"]{display:none!important}
[data-testid="stAppViewContainer"]>section:first-child{padding-top:56px!important}
#dpu-nb{position:fixed;top:0;left:0;right:0;height:48px;background:#1b3280;color:#fff;
  display:flex;align-items:center;padding:0 20px;gap:10px;z-index:999999;
  font-family:system-ui,sans-serif;font-size:14px;box-shadow:0 2px 12px rgba(0,0,0,.28)}
#dpu-nb a{color:#fff;text-decoration:none;opacity:.70}
#dpu-nb a:hover{opacity:1}
#dpu-nb .dm{width:26px;height:26px;background:#2e8cff;border-radius:6px;
  display:flex;align-items:center;justify-content:center;font-weight:800;font-size:11px}
#dpu-nb .sep{opacity:.25;font-size:18px}
#dpu-nb .dn{font-weight:600}
</style>
<div id="dpu-nb">
  <a href="https://calm-cocada-79e019.netlify.app/">← Hub</a>
  <span class="sep">|</span>
  <div class="dm">DE</div>
  <span class="dn">Generátor podkladů</span>
  <span id="dpu-usr"></span>
</div>
<script>
var u = new URLSearchParams(window.location.search).get("dpu_user");
if(u){ var el=document.getElementById("dpu-usr"); el.style.cssText="margin-left:auto;font-size:12px;opacity:.6"; el.textContent=u; }
</script>
""", unsafe_allow_html=True)
st.title("Podklady pro studie DPU ENERGY")
query = parse_query_params()

col1, col2, col3 = st.columns(3)
with col1:
    customer_name = st.text_input("Název zákazníka", value=query.get("customerName", ""))
with col2:
    project_code = st.text_input("Číslo projektu v Caflou", value=query.get("projectCode", ""))
with col3:
    project_name = st.text_input("Název projektu v Caflou", value=query.get("projectName", ""))

service_name = st.selectbox("Typ služby", list(SERVICES.keys()))
structure_mode_label = st.radio("Varianta struktury sdíleného uložiště", ["Neznáme rozsah objektů", "Máme seznam objektů"])
structure_mode = "object" if "seznam" in structure_mode_label.lower() else "standard"
service_df = build_service_dataframe(service_name)

object_input_df = pd.DataFrame(columns=OBJECT_OVERVIEW_HEADERS)
if structure_mode == "object":
    st.subheader("Seznam objektů pro objektovou variantu")
    manual_objects = st.text_area(
        "Pro zadání objektů vyplň na samostatné řádky název objektu a případně jeho adresu ve formátu Název objektu | Adresa s tím, že pro uložení je třeba stisknout na klávesnici kombinaci CTRL + ENTER.",
        height=140,
        placeholder="Název objektu A | Adresa objektu A\nNázev objektu B | Adresa objektu B"
    )
    object_input_df = parse_manual_objects(manual_objects)

    if not object_input_df.empty:
        st.dataframe(object_input_df, use_container_width=True, height=220)
    else:
        st.info("Zatím nemáš zadaný seznam objektů. Vygeneruje se šablona s prázdnými řádky.")

st.subheader("Seznam podkladů pro danou službu")
st.caption("Odškrtni položky, které po klientovi z nějakého důvodu nechceš požadovat.")

selected_rows = []
for block_no in sorted(service_df["block"].unique()):
    block_df = service_df[service_df["block"] == block_no].copy()
    block_name = BLOCK_DEFINITIONS[block_no]
    block_help = format_readme_for_ui(BLOCK_READMES.get(block_no, BLOCK_README_TEXTS.get(block_no, "")))
    with st.container(border=True):
        st.markdown(f"### {block_no}. {block_name}")
        intro_line = get_block_intro(block_no)
        if intro_line:
            st.caption(intro_line)
        if block_help:
            with st.expander("Co sem patří"):
                for i, row in block_df.iterrows():
                    label = f"{row['display_number']} | {row['name']} | {row['relevance_text']}"
                    checked = st.checkbox(label, value=True, key=f"chk_{i}")
                    if checked:
                        selected_rows.append(i)

selected_df = service_df.loc[selected_rows].copy()
if st.button("Zobrazit náhled checklistu"):
    if selected_df.empty:
        st.warning("Nemáš vybranou žádnou položku.")
    else:
        preview_df = selected_df[["display_number", "block_name", "name", "description", "relevance_text"]].rename(columns={"display_number": "Číslo položky", "block_name": "Blok", "name": "Položka", "description": "Co se očekává", "relevance_text": "Relevance"})
        st.dataframe(preview_df, use_container_width=True, height=650)

if selected_df.empty:
    st.warning("Vyber alespoň jednu položku seznamu.")
    st.stop()

checklist_bytes = build_checklist_xlsx(selected_df, service_name)
prehled_objektu_bytes = build_prehled_objektu_xlsx(selected_df, object_input_df, object_mode=(structure_mode == "object"))
centralni_prehled_bytes = build_centralni_prehled_xlsx(selected_df, service_name) if structure_mode == "object" else None
zip_bytes = build_zip_package(customer_name=customer_name, project_code=project_code, project_name=project_name, service_name=service_name, checklist_bytes=checklist_bytes, prehled_objektu_bytes=prehled_objektu_bytes, selected_df=selected_df, structure_mode=structure_mode, centralni_prehled_bytes=centralni_prehled_bytes, object_df=object_input_df)

st.subheader("Stažení výstupů")
st.download_button("Stáhnout CHECKLIST_PODKLADU.xlsx", data=checklist_bytes, file_name=f"CHECKLIST_PODKLADU_{safe_name(project_code)}_{safe_name(SERVICES[service_name]['short'])}.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
st.download_button("Stáhnout PREHLED_OBJEKTU.xlsx", data=prehled_objektu_bytes, file_name=f"PREHLED_OBJEKTU_{safe_name(project_code)}.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
if centralni_prehled_bytes is not None:
    st.download_button("Stáhnout PREHLED_CENTRALNICH_PODKLADU.xlsx", data=centralni_prehled_bytes, file_name=f"PREHLED_CENTRALNICH_PODKLADU_{safe_name(project_code)}.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
st.download_button("Stáhnout ZIP se složkovou strukturou", data=zip_bytes, file_name=f"{get_service_root_folder(service_name)}.zip", mime="application/zip")
